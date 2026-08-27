"""Parsing a class list out of whatever the lecturer actually uploads.

The spreadsheet a Ghanaian department hands round is not a schema. Across the
files we have seen the email column is called "Email", "E-mail", "Email
Address", "Student Email" or "mail", the name arrives either split into two
columns or as one "Name"/"Full Name" column, and there are usually a few blank
rows and a title row above the header.

So this module does not ask for a template. It finds the header row by looking
for something that resembles an email column, maps the columns it recognises,
and reports every row it could not use rather than failing the whole import —
one malformed address in a list of two hundred must not cost the lecturer the
other hundred and ninety-nine.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field

# Deliberately permissive. This is a sanity check that catches "N/A", a phone
# number or a truncated address; the authoritative check is whether Supabase
# can send to it. Rejecting valid-but-unusual addresses would silently drop
# students from a roster, which is the failure that matters here.
_EMAIL_RE = re.compile(r"^[^@\s,;]+@[^@\s,;]+\.[^@\s,;]{2,}$")

# Header aliases, lowercased and stripped of punctuation.
_EMAIL_HEADERS = {
    "email", "emailaddress", "email address", "e mail", "e mail address",
    "mail", "student email", "studentemail", "student mail", "institutional email",
    "school email", "university email",
}
_FIRST_HEADERS = {
    "first name", "firstname", "first", "given name", "givenname", "forename",
    "other names", "othernames",
}
_LAST_HEADERS = {
    "last name", "lastname", "last", "surname", "family name", "familyname",
}
_FULL_HEADERS = {
    "name", "full name", "fullname", "student name", "studentname",
    "student", "names",
}

MAX_ROWS = 2000


@dataclass
class RosterRow:
    email: str
    first_name: str | None = None
    last_name: str | None = None


@dataclass
class RosterParse:
    rows: list[RosterRow] = field(default_factory=list)
    # (row number as the lecturer sees it in the spreadsheet, reason)
    skipped: list[tuple[int, str]] = field(default_factory=list)


class RosterFormatError(ValueError):
    """The file could not be read at all, or carries no email column."""


def _norm(value: object) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", str(value or "").strip().lower()).strip()


def _split_full_name(full: str) -> tuple[str | None, str | None]:
    """"Kwame Nkrumah" -> ("Kwame", "Nkrumah").

    Rosters are also written surname-first as "Nkrumah, Kwame"; that comma is
    the only reliable signal of the ordering, so it is the only one used.
    """
    full = re.sub(r"\s+", " ", (full or "").strip())
    if not full:
        return None, None
    if "," in full:
        last, _, first = full.partition(",")
        return first.strip() or None, last.strip() or None
    parts = full.split(" ")
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def _cells_to_text(row: list) -> list[str]:
    return ["" if c is None else str(c).strip() for c in row]


def _find_header(grid: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Locate the header row and map it to column indices.

    Scans the first 20 rows rather than assuming row 1: exported class lists
    routinely carry a course title and a blank line above the real header.
    """
    for index, row in enumerate(grid[:20]):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _norm(cell)
            if not key:
                continue
            if key in _EMAIL_HEADERS and "email" not in mapping:
                mapping["email"] = col
            elif key in _FIRST_HEADERS and "first" not in mapping:
                mapping["first"] = col
            elif key in _LAST_HEADERS and "last" not in mapping:
                mapping["last"] = col
            elif key in _FULL_HEADERS and "full" not in mapping:
                mapping["full"] = col
        if "email" in mapping:
            return index, mapping

    # No recognisable header. If some column is plainly full of addresses the
    # file is still usable — a headerless two-column export is common.
    for index, row in enumerate(grid[:20]):
        for col, cell in enumerate(row):
            if _EMAIL_RE.match(cell.strip()):
                return index - 1, {"email": col, "full": 0 if col != 0 else 1}

    raise RosterFormatError(
        "No email column found. The sheet needs a column headed 'Email' "
        "(a 'Name' or 'First name'/'Last name' column is optional)."
    )


def _parse_grid(grid: list[list[str]]) -> RosterParse:
    header_index, mapping = _find_header(grid)
    out = RosterParse()
    seen: set[str] = set()

    for offset, row in enumerate(grid[header_index + 1:], start=header_index + 2):
        if len(out.rows) >= MAX_ROWS:
            out.skipped.append((offset, f"Import is capped at {MAX_ROWS} students per file"))
            break

        def cell(key: str) -> str:
            col = mapping.get(key)
            return row[col].strip() if col is not None and col < len(row) else ""

        email = cell("email")
        if not email and not any(c.strip() for c in row):
            continue  # blank spacer row — not worth reporting
        if not email:
            out.skipped.append((offset, "No email address"))
            continue
        if not _EMAIL_RE.match(email):
            out.skipped.append((offset, f"{email!r} is not a valid email address"))
            continue

        key = email.lower()
        if key in seen:
            out.skipped.append((offset, f"{email} appears more than once in this file"))
            continue
        seen.add(key)

        first = cell("first") or None
        last = cell("last") or None
        if not first and not last:
            first, last = _split_full_name(cell("full"))

        out.rows.append(RosterRow(email=email, first_name=first, last_name=last))

    return out


def _read_csv(content: bytes) -> list[list[str]]:
    # utf-8-sig, not utf-8: Excel's "CSV UTF-8" export writes a BOM, and the
    # BOM would otherwise become part of the first header cell and stop the
    # email column being recognised.
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = content.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return [_cells_to_text(r) for r in csv.reader(io.StringIO(text), dialect)]

    raise RosterFormatError("The CSV file could not be decoded as text.")


def _read_xlsx(content: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — dependency is pinned
        raise RosterFormatError(
            "Excel files need the openpyxl package on the server. "
            "Save the sheet as CSV and upload that instead."
        ) from exc

    try:
        # read_only keeps a 2000-row sheet off the heap; data_only returns the
        # cached value of a formula rather than "=CONCAT(A2,B2)".
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises several types
        raise RosterFormatError(f"The spreadsheet could not be opened: {exc}") from exc

    try:
        sheet = workbook.active
        return [_cells_to_text(list(row)) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_roster(content: bytes, filename: str) -> RosterParse:
    """Parse an uploaded roster. Raises RosterFormatError if unreadable."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xlsm"):
        grid = _read_xlsx(content)
    elif ext in ("csv", "txt", "tsv"):
        grid = _read_csv(content)
    elif ext == "xls":
        raise RosterFormatError(
            "The old .xls format is not supported. Open it in Excel and save "
            "as .xlsx or .csv."
        )
    else:
        raise RosterFormatError(
            f"Unsupported file type '.{ext}'. Upload a .csv or .xlsx file."
        )

    if not grid:
        raise RosterFormatError("The file is empty.")

    return _parse_grid(grid)
